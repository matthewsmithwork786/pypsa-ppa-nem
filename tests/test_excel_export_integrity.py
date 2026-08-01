"""W11 regression: no plain note/label string may be stored as an Excel formula.

openpyxl treats any string value starting with ``=`` as a formula. A note like
``= FID; devex bullet and construction both start here`` therefore lands in
``xl/worksheets/sheet2.xml`` as an unparseable ``<f>`` element, and Excel reports
"Removed Records: Formula from /xl/worksheets/sheet2.xml part" on open.

The fix (W11) removes the leading ``=`` and adds a ``_text()`` guard so every
label/note/unit is written as literal text. These tests assert that every
intentionally-formula cell is syntactically plausible and that no note/label
position ever carries a formula.
"""
from __future__ import annotations

import re
import zipfile
from io import BytesIO
from xml.etree import ElementTree as ET

import openpyxl
import pytest

from ppa.financial_model import EnergyInputs, ProjectFinanceInputs, run_project_finance
from ppa.financial_model_excel import export_financial_model

# A syntactically-plausible Excel formula's first non-space character must be one
# of these (openpyxl emits formulas without the leading '=').
_PLAUSIBLE_FORMULA_START = set("ABCDEFGHIJKLMNOPQRSTUVWXYZ(+-@'0-9")

# The exact note string that caused the corruption (leading '=' before "FID").
_BUGGY_DEVELOPMENT_START_NOTE = "= FID; devex bullet and construction both start here"


def _simple_energy() -> EnergyInputs:
    return EnergyInputs(
        onsw_mw=100.0,
        pv_mw=0.0,
        bess_mw=0.0,
        bess_mwh=0.0,
        load_mw=50.0,
        ppa_gwh=300.0,
        excess_solar_gwh=0.0,
        excess_nonsolar_gwh=50.0,
        penalty_gwh=5.0,
        total_solar_gwh=0.0,
        total_nonsolar_gwh=350.0,
        sell_solar_price=60.0,
        sell_nonsolar_price=55.0,
        purchase_price=70.0,
        marketbuy_gwh=2.0,
        name="excel-integrity test",
    )


def _export_bytes() -> bytes:
    p = ProjectFinanceInputs()
    e = _simple_energy()
    result = run_project_finance(p, e)
    return export_financial_model(p, e, result)


def _sheet_xml_files(xlsx_bytes: bytes) -> dict[str, str]:
    """Map worksheet name -> raw sheetN.xml string (resolving via rels)."""
    with zipfile.ZipFile(BytesIO(xlsx_bytes)) as z:
        wb_xml = z.read("xl/workbook.xml").decode("utf-8")
        rels_xml = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")

    ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
          "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships"}
    sheet_targets: dict[str, str] = {}
    wb_root = ET.fromstring(wb_xml)
    for sheet in wb_root.findall(".//m:sheets/m:sheet", ns):
        rid = sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
        sheet_targets[sheet.attrib["name"]] = rid

    rel_root = ET.fromstring(rels_xml)
    rel_map: dict[str, str] = {}
    for rel in rel_root.findall("{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"):
        rel_map[rel.attrib["Id"]] = rel.attrib["Target"]

    result: dict[str, str] = {}
    with zipfile.ZipFile(BytesIO(xlsx_bytes)) as z:
        for name, rid in sheet_targets.items():
            target = rel_map[rid]
            if not target.startswith("xl/"):
                target = target[1:] if target.startswith("/xl/") else "xl/" + target.lstrip("/")
            result[name] = z.read(target).decode("utf-8")
    return result


# ── Raw-XML regression: the Inputs sheet must contain zero <f> elements ──────

@pytest.mark.xfail(strict=True, reason="W11: note still written with a leading '='")
def test_inputs_sheet_contains_no_formula_elements():
    """The Inputs sheet has no real formulas, so any `<f>` element in its XML is
    a note/label accidentally stored as a formula. W11 must leave it empty."""
    sheet_xmls = _sheet_xml_files(_export_bytes())
    inputs_xml = sheet_xmls["Inputs"]
    formulas = re.findall(r"<f>(.*?)</f>", inputs_xml, re.DOTALL)
    assert formulas == [], (
        f"Inputs sheet contains {len(formulas)} spurious formula element(s): {formulas[:3]}"
    )


def test_every_formula_element_is_plausible():
    """Every `<f>` across every worksheet must look like a real formula: its
    first non-space character must be a formula-leading token."""
    sheet_xmls = _sheet_xml_files(_export_bytes())
    assert sheet_xmls, "expected at least one worksheet"
    for sheet_name, xml in sheet_xmls.items():
        for match in re.findall(r"<f>(.*?)</f>", xml, re.DOTALL):
            stripped = match.strip()
            assert stripped, f"empty <f> element on sheet {sheet_name!r}"
            assert stripped[0] in _PLAUSIBLE_FORMULA_START, (
                f"unplausible formula {match!r} on sheet {sheet_name!r}"
            )


# ── openpyxl-level regression on the exact bug ───────────────────────────────

@pytest.mark.xfail(strict=True, reason="W11: development_start note is written as a formula")
def test_development_start_note_is_literal_text():
    """The 'Development start period' note must be a plain string cell
    (data_type 's'), never a formula, and its text must not start with '='."""
    wb = openpyxl.load_workbook(BytesIO(_export_bytes()), data_only=False)
    inputs_ws = wb["Inputs"]
    found = None
    for row in inputs_ws.iter_rows():
        if row[1].value == "Development start period":
            found = row[4]  # column E holds the note
            break
    assert found is not None, "could not find 'Development start period' note cell"
    assert found.data_type != "f", (
        f"note stored as a formula: {found.value!r} at {found.coordinate}"
    )
    assert isinstance(found.value, str) and not found.value.startswith("="), (
        f"note text must not start with '=': {found.value!r}"
    )
    assert found.value != _BUGGY_DEVELOPMENT_START_NOTE


def test_notes_sheet_is_all_literal_text():
    """The Notes sheet is prose only; no cell there may be stored as a formula."""
    wb = openpyxl.load_workbook(BytesIO(_export_bytes()), data_only=False)
    notes_ws = wb["Notes"]
    for row in notes_ws.iter_rows():
        for cell in row:
            if cell.value is not None and cell.data_type == "f":
                raise AssertionError(
                    f"Notes sheet cell {cell.coordinate} stored as a formula: {cell.value!r}"
                )


def test_model_units_column_is_never_formula():
    """The Model sheet's unit column (C) holds 'A$M'/etc. — a 'A$...' text must
    not be misread as a formula (openpyxl would need no leading '=', but guard
    that no such cell is marked as a formula)."""
    wb = openpyxl.load_workbook(BytesIO(_export_bytes()), data_only=False)
    model_ws = wb["Model"]
    for row in model_ws.iter_rows():
        cell = row[2] if len(row) > 2 else None
        if cell is not None and cell.value is not None:
            assert cell.data_type != "f", (
                f"unit/label cell {cell.coordinate} stored as a formula: {cell.value!r}"
            )
