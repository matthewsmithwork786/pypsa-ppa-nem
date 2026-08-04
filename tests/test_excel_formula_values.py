"""Cross-check: the exported workbook's live formulas evaluate to the same
numbers as ``ppa/financial_model.py``.

Uses the ``formulas`` Excel-evaluator package (skipped when not installed) to
recalculate the whole workbook, then asserts the Model-sheet schedule rows and
the Outputs-sheet KPIs agree with :func:`run_project_finance`. Tolerance is a
bit loose (1e-4) because debt drawdown/IDC/interest are stored rounded to 6 d.p.
and the cumulative rows roll that rounding up.
"""
from __future__ import annotations

from io import BytesIO

import numpy as np
import pytest
import openpyxl

formulas = pytest.importorskip("formulas")

from ppa.financial_model import EnergyInputs, ProjectFinanceInputs, run_project_finance
from ppa.financial_model_excel import export_financial_model


def _simple_energy() -> EnergyInputs:
    return EnergyInputs(
        onsw_mw=100.0,
        pv_mw=0.0,
        bess_mw=0.0,
        bess_mwh=0.0,
        load_mw=50.0,
        ppa_gwh=300.0,
        excess_solar_gwh=0.0,
        excess_nonsolar_gwh=50.0,
        penalty_gwh=5.0,
        total_solar_gwh=0.0,
        total_nonsolar_gwh=350.0,
        sell_solar_price=60.0,
        sell_nonsolar_price=55.0,
        purchase_price=70.0,
        marketbuy_gwh=2.0,
        name="formula-values test",
    )


# Excel Model-sheet row label -> schedule key in ProjectFinanceResult.schedule.
_MODEL_LABELS = {
    "Devex": "devex",
    "Capex": "capex",
    "Total capital spend": "total_capital_spend",
    "PPA delivered volume": "ppa_vol",
    "Merchant volume — solar hours": "merch_solar_vol",
    "Merchant volume — non-solar hours": "merch_nonsolar_vol",
    "Penalty generation volume": "pen_vol",
    "Average PPA price": "ppa_price",
    "Average merchant price — solar hours": "merch_solar_price",
    "Average merchant price — non-solar hours": "merch_nonsolar_price",
    "Average penalty price": "pen_price",
    "LGC / GO price": "lgc_price",
    "PPA revenue": "ppa_rev",
    "Merchant revenue — solar hours": "merch_solar_rev",
    "Merchant revenue — non-solar hours": "merch_nonsolar_rev",
    "Penalty cost": "penalty_cost",
    "Total revenue": "total_rev",
    "Total O&M expenses": "opex",
    "EBITDA": "ebitda",
    "Opening balance": "debt_opening",
    "Closing balance": "debt_closing",
    "Book depreciation": "book_dep",
    "Tax depreciation": "tax_dep",
    "Profit before tax": "pbt",
    "Income tax": "tax",
    "Profit after tax": "pat",
    "Cash flow from operations": "ocf",
    "Cash flow from investing": "investing_cf",
    "Cash flow from financing": "financing_cf",
    "Net cash flow": "net_cashflow",
    "Cumulative net cash flow": "cum_net_cf",
    "Fixed assets (gross, incl. capitalised IDC)": "gross_fixed_assets",
    "Accumulated depreciation": "accum_dep",
    "Net fixed assets": "net_fixed_assets",
    "Net debt": "net_debt",
    "Shareholders' funds": "net_equity",
    "Equity investment": "equity_spend",
    "FCFF (project)": "fcff",
    "FCFE (equity)": "fcfe",
    "Cumulative FCFE": "cum_fcfe",
    "CFADS": "cfads",
    "DSCR": "dscr",
}

_OUTPUT_LABELS = {
    "Project IRR (FCFF)": "project_irr",
    "Equity IRR (FCFE)": "equity_irr",
    "NPV @ WACC (project)": "npv_project",
    "Gearing": "gearing",
    "Total debt": "total_debt",
    "Total equity": "total_equity",
    "Minimum DSCR": "min_dscr",
    "Average DSCR": "avg_dscr",
    "Equity payback": "payback_years",
    "LCOE": "lcoe",
}

_COL_D = 4  # column D holds year 0


def _as_float(value) -> float:
    if value is None:
        return float("nan")
    try:
        return float(np.asarray(value, dtype=float).reshape(-1)[0])
    except (TypeError, ValueError):
        pass
    inner = getattr(value, "value", value)
    try:
        return float(np.asarray(inner, dtype=float).reshape(-1)[0])
    except (TypeError, ValueError):
        return float("nan")


def _eval_workbook(tmp_path, xlsx_bytes: bytes):
    """Recalculate *xlsx_bytes* and return {title: {cell: value}} by coordinate."""
    path = tmp_path / "financial-model.xlsx"
    path.write_bytes(xlsx_bytes)
    model = formulas.ExcelModel().loads(str(path)).finish()
    solution = model.calculate()
    out: dict[str, dict[str, object]] = {}
    for ref, value in solution.items():
        # Ref shape: '[<file>]SHEET'!<cell>; sheet names come back upper-cased.
        sheet = ref.lstrip("'").split("'!")[0].split("]")[-1]
        cell = ref.split("'!")[1]
        out.setdefault(sheet, {})[cell] = value
    return out


def _by_title(sheets: dict[str, dict[str, object]], title: str):
    """Look up a sheet by openpyxl title (formulas uppercases sheet names)."""
    return sheets.get(title.upper(), sheets.get(title, {}))


@pytest.fixture(scope="module")
def workbook_result():
    p = ProjectFinanceInputs()
    e = _simple_energy()
    return p, e, run_project_finance(p, e)


def test_model_rows_evaluate_to_python_schedule(workbook_result, tmp_path):
    p, e, result = workbook_result
    xlsx_bytes = export_financial_model(p, e, result)
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=False)
    sheets = _eval_workbook(tmp_path, xlsx_bytes)
    model_sol = _by_title(sheets, "Model")
    model_ws = wb["Model"]

    # Map row labels -> row numbers.
    row_of = {}
    for row in model_ws.iter_rows(min_col=2, max_col=2):
        if row[0].value in _MODEL_LABELS:
            row_of[row[0].value] = row[0].row

    schedule = result.schedule
    mismatches = []
    for label, key in _MODEL_LABELS.items():
        r = row_of.get(label)
        assert r is not None, f"Model row {label!r} not found"
        expected = schedule[key]
        for year in range(p.model_duration):
            col = openpyxl.utils.get_column_letter(_COL_D + year)
            got = _as_float(model_sol.get(f"{col}{r}"))
            want = float(expected[year])
            if not (np.isclose(got, want, atol=1e-4) or
                    (np.isnan(got) and np.isnan(want))):
                mismatches.append((label, year, got, want))
    assert not mismatches, (
        "Excel-evaluated Model rows disagree with Python schedule:\n"
        + "\n".join(f"  {lab} year {yr}: xl={got!r} py={want!r}" for lab, yr, got, want in mismatches[:20])
    )


def test_outputs_kpis_evaluate_to_python(workbook_result, tmp_path):
    p, e, result = workbook_result
    xlsx_bytes = export_financial_model(p, e, result)
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=False)
    sheets = _eval_workbook(tmp_path, xlsx_bytes)
    outputs_sol = _by_title(sheets, "Outputs")
    outputs_ws = wb["Outputs"]

    row_of = {}
    for row in outputs_ws.iter_rows(min_col=2, max_col=2):
        if row[0].value in _OUTPUT_LABELS:
            row_of[row[0].value] = row[0].row

    mismatches = []
    for label, attr in _OUTPUT_LABELS.items():
        r = row_of.get(label)
        assert r is not None, f"Outputs KPI {label!r} not found"
        got = _as_float(outputs_sol.get(f"C{r}"))
        want = float(getattr(result, attr))
        if not (np.isclose(got, want, atol=1e-4) or
                (np.isnan(got) and np.isnan(want))):
            mismatches.append((label, got, want))
    assert not mismatches, (
        "Excel-evaluated Outputs KPIs disagree with Python:\n"
        + "\n".join(f"  {lab}: xl={got!r} py={want!r}" for lab, got, want in mismatches)
    )
