"""The hourly timeseries ships as its own workbook; the numbers must still agree.

25 years x 8760 rows made the financial model slow to build and heavy to
download, so the hours moved into a separate file and the Energy tab now
carries hard values. That trade is only acceptable if the hard values are
exactly what the hourly workbook sums to — otherwise the two downloads tell
different stories about the same run.
"""
from __future__ import annotations

import numpy as np
import openpyxl
import pandas as pd
import pytest

from ppa.financial_model import (
    EnergyInputs,
    ProjectFinanceInputs,
    energy_inputs_from_results,
    run_project_finance,
)
from ppa.financial_model_excel import export_financial_model, export_hourly_timeseries

from io import BytesIO


# ── Minimal stand-ins for a per-year simulation result ───────────────────────

class _Dispatch:
    def __init__(self, n, seed):
        rng = np.random.default_rng(seed)
        idx = pd.date_range("2025-01-01", periods=n, freq="h")
        hour = idx.hour.to_numpy()
        solar_shape = np.clip(np.sin((hour - 6) / 12 * np.pi), 0, None)
        self.wind_gen = pd.Series(20 + 10 * rng.random(n), index=idx)
        self.pv_gen = pd.Series(60 * solar_shape, index=idx)
        self.bess_dispatch = pd.Series(5 * rng.random(n), index=idx)
        self.bess_store = pd.Series(4 * rng.random(n), index=idx)
        self.market_buy = pd.Series(2 * rng.random(n), index=idx)
        self.market_sell = pd.Series(8 * rng.random(n), index=idx)
        self.ppa_delivery = pd.Series(30 + 5 * rng.random(n), index=idx)
        self.penalty_gen = pd.Series(3 * rng.random(n), index=idx)
        self.index = idx


class _Scenario:
    first_sim_year = 2025
    name = "split test"
    onsw_mw = 100.0
    pv_mw = 80.0
    bess_mw = 10.0
    bess_mwh = 40.0
    effective_bess_mw = 10.0
    effective_bess_mwh = 40.0
    ppaload_mw = 50.0


class _Summary:
    def __init__(self, d):
        self.ppa_delivered_mwh = float(d.ppa_delivery.sum())
        self.penalty_mwh = float(d.penalty_gen.sum())


class _YearResult:
    def __init__(self, n, seed):
        self.dispatch = _Dispatch(n, seed)
        self.market_prices = pd.Series(
            50 + 40 * np.random.default_rng(seed + 99).random(n), index=self.dispatch.index
        )
        self.scenario = _Scenario()
        self.summary = _Summary(self.dispatch)
        self.n_period_hours = n


@pytest.fixture(scope="module")
def year_results():
    # Two short years keeps the test quick; the roll-up maths is per-year, so
    # length does not change what is being asserted.
    return [_YearResult(168, seed) for seed in (1, 2)]


def _sheet_values(xlsx: bytes, sheet: str) -> openpyxl.worksheet.worksheet.Worksheet:
    return openpyxl.load_workbook(BytesIO(xlsx), data_only=False)[sheet]


def test_financial_model_has_no_hourly_sheet(year_results):
    p = ProjectFinanceInputs()
    e = energy_inputs_from_results(year_results)
    xlsx = export_financial_model(p, e, run_project_finance(p, e))
    wb = openpyxl.load_workbook(BytesIO(xlsx))
    assert "Hourly" not in wb.sheetnames, (
        "the hours must not ride along in the finance workbook any more"
    )
    assert {"Outputs", "Inputs", "Energy", "Model", "Notes"} <= set(wb.sheetnames)


def test_hourly_workbook_holds_every_hour_of_every_year(year_results):
    xlsx = export_hourly_timeseries(year_results)
    wb = openpyxl.load_workbook(BytesIO(xlsx))
    assert wb.sheetnames == ["Hourly"]
    ws = wb["Hourly"]
    years = [ws.cell(r, 1).value for r in range(17, ws.max_row + 1)]
    years = [y for y in years if y is not None]
    assert len(years) == sum(len(r.dispatch.wind_gen) for r in year_results)
    assert sorted(set(years)) == [2025, 2026]


def test_energy_tab_values_are_hard_numbers_not_formulas(year_results):
    p = ProjectFinanceInputs()
    e = energy_inputs_from_results(year_results)
    xlsx = export_financial_model(p, e, run_project_finance(p, e))
    ws = _sheet_values(xlsx, "Energy")
    formulas = [
        (ws.cell(r, 2).value, ws.cell(r, 3).value)
        for r in range(4, 30)
        if isinstance(ws.cell(r, 3).value, str) and str(ws.cell(r, 3).value).startswith("=")
    ]
    assert not formulas, f"Energy tab still references the Hourly sheet: {formulas}"


def test_energy_hard_values_match_the_hourly_workbook(year_results):
    """The whole point of the split: both downloads must describe one run.

    Recomputes each Energy figure straight from the rows written into the hourly
    workbook and compares against the hard value on the Energy tab.
    """
    p = ProjectFinanceInputs()
    e = energy_inputs_from_results(year_results)
    fm = export_financial_model(p, e, run_project_finance(p, e))
    hourly = export_hourly_timeseries(year_results)

    ws = openpyxl.load_workbook(BytesIO(hourly))["Hourly"]
    rows = [[ws.cell(r, c).value for c in range(1, 14)] for r in range(17, ws.max_row + 1)]
    df = pd.DataFrame(
        rows,
        columns=["year", "ts", "hour", "wind", "pv", "bess_dis", "bess_chg",
                 "total", "buy", "sell", "ppa", "penalty", "price"],
    ).dropna(subset=["year"])

    solar_hours = (df["hour"] >= 9) & (df["hour"] < 17)
    per_year = []
    for year, grp in df.groupby("year"):
        scale = 8760 / len(grp)  # the sheet's annualisation factor
        mask = solar_hours.loc[grp.index]
        per_year.append({
            "ppa_gwh": grp["ppa"].sum() * scale / 1000,
            "excess_solar_gwh": grp.loc[mask, "sell"].sum() * scale / 1000,
            "penalty_gwh": grp["penalty"].sum() * scale / 1000,
            "total_solar_gwh": grp.loc[mask, "total"].sum() * scale / 1000,
            "marketbuy_gwh": grp["buy"].sum() * scale / 1000,
        })
    expected = pd.DataFrame(per_year).mean().to_dict()

    energy_ws = openpyxl.load_workbook(BytesIO(fm), data_only=False)["Energy"]
    labels = {
        "PPA delivered": "ppa_gwh",
        "Excess sold — solar hours": "excess_solar_gwh",
        "Penalty (undelivered)": "penalty_gwh",
        "Total generation — solar hours": "total_solar_gwh",
        "Market purchase volume": "marketbuy_gwh",
    }
    seen = {}
    for r in range(4, 30):
        label = energy_ws.cell(r, 2).value
        if label in labels:
            seen[labels[label]] = float(energy_ws.cell(r, 3).value)

    assert set(seen) == set(labels.values()), f"missing Energy rows: {set(labels.values()) - set(seen)}"
    mismatches = {
        k: (seen[k], expected[k])
        for k in expected
        # Hourly cells are rounded to 3 d.p. on write, so allow a hair of drift.
        if not np.isclose(seen[k], expected[k], rtol=1e-4, atol=1e-6)
    }
    assert not mismatches, (
        "Energy tab disagrees with the hourly workbook (value, recomputed): "
        f"{mismatches}"
    )
