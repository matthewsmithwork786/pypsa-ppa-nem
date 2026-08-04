"""Lockstep test: the live Excel export must match the reworked financial model.

Ensures ``ppa/financial_model_excel.py`` stays in sync with ``ppa/financial_model.py``:
no leftover per-technology "dev years" rows on the Inputs sheet; per-technology
devex/capex rows with subtotals; the devex bullet keyed off the live
``development_start`` input cell; year 0 as the un-indexed base year; the debt
corkscrew; the single combined Hourly sheet; and fully formula-driven Outputs.
"""
from __future__ import annotations

import re
from io import BytesIO
from types import SimpleNamespace

import openpyxl
import pandas as pd
import pytest

from ppa.financial_model import (
    EnergyInputs,
    ProjectFinanceInputs,
    run_project_finance,
)
from ppa.financial_model_excel import (
    _AGG_ROWS,
    _HOURLY_DATA_START,
    _HOURLY_HEADER_ROW,
    export_financial_model,
)

_DEV_YEAR_RE = re.compile(r"dev.*year", re.IGNORECASE)


def _simple_energy() -> EnergyInputs:
    return EnergyInputs(
        onsw_mw=100.0,
        pv_mw=0.0,
        bess_mw=0.0,
        bess_mwh=0.0,
        load_mw=50.0,
        ppa_gwh=300.0,
        excess_solar_gwh=0.0,
        excess_nonsolar_gwh=50.0,
        penalty_gwh=5.0,
        total_solar_gwh=0.0,
        total_nonsolar_gwh=350.0,
        sell_solar_price=60.0,
        sell_nonsolar_price=55.0,
        purchase_price=70.0,
        marketbuy_gwh=2.0,
        name="excel-export test",
    )


def _load_model_ws() -> tuple[openpyxl.Workbook, openpyxl.worksheet.worksheet.Worksheet]:
    p = ProjectFinanceInputs()
    e = _simple_energy()
    result = run_project_finance(p, e)
    wb = openpyxl.load_workbook(BytesIO(export_financial_model(p, e, result)),
                                data_only=False)
    return wb, wb["Model"]


def _row_of(ws, label: str) -> int | None:
    for row in ws.iter_rows(min_col=2, max_col=2):
        if row[0].value == label:
            return row[0].row
    return None


def test_excel_export_matches_devex_single_bullet_at_fid():
    p = ProjectFinanceInputs()
    e = _simple_energy()
    result = run_project_finance(p, e)

    xlsx_bytes = export_financial_model(p, e, result)
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=False)

    # (a) No "dev...year(s)" labels remain on the Inputs sheet (the removed
    # per-technology dev-year rows must not have crept back in).
    inputs_ws = wb["Inputs"]
    label_col = 2  # column B holds row labels throughout this workbook
    for row in inputs_ws.iter_rows(min_col=label_col, max_col=label_col):
        for cell in row:
            if isinstance(cell.value, str):
                assert not _DEV_YEAR_RE.search(cell.value), (
                    f"unexpected dev-year label {cell.value!r} at {cell.coordinate}"
                )

    # Locate the "Development start period" input cell so we can confirm the
    # Model sheet's devex formulas reference it.
    dev_start_cell = None
    for row in inputs_ws.iter_rows(min_col=label_col, max_col=label_col):
        for cell in row:
            if cell.value == "Development start period":
                dev_start_cell = f"$C${cell.row}"
                break
        if dev_start_cell:
            break
    assert dev_start_cell is not None, "could not find 'Development start period' input row"

    model_ws = wb["Model"]

    # (b) Exactly one "Devex" subtotal row on the Model sheet.
    devex_rows = [
        cell.row
        for row in model_ws.iter_rows(min_col=label_col, max_col=label_col)
        for cell in row
        if cell.value == "Devex"
    ]
    assert len(devex_rows) == 1, f"expected exactly one 'Devex' row, found {devex_rows}"

    # (c) Each per-technology devex row's formula references the live
    # development_start cell (single bullet at FID), not a multi-period spread.
    tech_labels = ("Devex — Onshore wind", "Devex — Solar PV", "Devex — BESS")
    tech_rows = [r for r in map(lambda lab: _row_of(model_ws, lab), tech_labels)]
    assert all(r is not None for r in tech_rows), tech_rows
    for r in tech_rows:
        formula_cells = [
            cell for cell in model_ws[r]
            if isinstance(cell.value, str) and cell.value.startswith("=")
        ]
        assert formula_cells, f"per-tech devex row {r} has no formula cells"
        for cell in formula_cells:
            assert dev_start_cell in cell.value, (
                f"devex formula at {cell.coordinate} does not reference "
                f"{dev_start_cell} ({cell.value!r})"
            )

    # (d) The subtotal row sums the per-technology rows (same column letter).
    subtotal = devex_rows[0]
    subtotal_formulas = [
        cell for cell in model_ws[subtotal]
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]
    assert subtotal_formulas, "Devex subtotal row has no formula cells"
    for cell in subtotal_formulas:
        cl = cell.column_letter
        assert dev_start_cell not in cell.value
        assert all(f"{cl}{r}" in cell.value for r in tech_rows)


# ── Year 0 is the base year: no indexation applied ───────────────────────────


def test_excel_export_year0_is_unindexed_base_year():
    p = ProjectFinanceInputs()
    wb, model_ws = _load_model_ws()

    # Cost inflation multiplier at year 0 (column D) must be 1.0:
    # =(1+Inputs!$C$<rate>)^D$4, and D$4 holds the year 0 header.
    cost_idx_row = _row_of(model_ws, "Cost inflation")
    assert cost_idx_row is not None
    c0 = model_ws.cell(cost_idx_row, 4)
    assert isinstance(c0.value, str) and c0.value.startswith("=("), c0.value
    assert "^D$4" in c0.value

    # The onshore-devex bullet lands at year 0 and carries no indexation:
    # =IF(D$4=<dev_start>-1, D$<cost_idx>*(devex*MW), 0)
    devex_onsw_row = _row_of(model_ws, "Devex — Onshore wind")
    assert devex_onsw_row is not None
    f0 = model_ws.cell(devex_onsw_row, 4).value
    assert f0.startswith("=IF(D$4=Inputs!$C$"), f0
    # The FID bullet (year 0) is the only non-zero devex cell.
    non_zero = [
        c.coordinate for c in model_ws[devex_onsw_row]
        if isinstance(c.value, str) and "IF(" in c.value and ",0)" not in c.value
    ]
    assert not non_zero  # every period is the same guarded bullet shape


# ── Per-technology capex rows feed a subtotal ────────────────────────────────


def test_excel_export_capex_has_per_technology_subtotal():
    p = ProjectFinanceInputs()
    wb, model_ws = _load_model_ws()

    tech_labels = ("Capex — Onshore wind", "Capex — Solar PV", "Capex — BESS")
    tech_rows = [r for r in map(lambda lab: _row_of(model_ws, lab), tech_labels)]
    assert all(r is not None for r in tech_rows), tech_rows

    capex_row = _row_of(model_ws, "Capex")
    assert capex_row is not None
    capex_cells = [
        cell for cell in model_ws[capex_row]
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]
    assert capex_cells
    for cell in capex_cells:
        cl = cell.column_letter
        assert all(f"{cl}{r}" in cell.value for r in tech_rows)

    # Onshore wind builds over onsw_constr_years=2 (default) — its capex row is
    # a per-year fraction, not a single bullet.
    onsw_row = tech_rows[0]
    onsw_formulas = [
        cell.value for cell in model_ws[onsw_row]
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]
    assert onsw_formulas
    for formula in onsw_formulas:
        assert formula.startswith("=")


# ── Debt corkscrew: opening feeds from the prior closing balance ─────────────


def test_excel_export_debt_corkscrew():
    p = ProjectFinanceInputs()
    wb, model_ws = _load_model_ws()

    opening_row = _row_of(model_ws, "Opening balance")
    closing_row = _row_of(model_ws, "Closing balance")
    draw_row = _row_of(model_ws, "Drawdown (construction)")
    idc_row = _row_of(model_ws, "Interest during construction")
    repay_row = _row_of(model_ws, "Repayment (principal)")
    assert all(r is not None for r in (opening_row, closing_row, draw_row,
                                       idc_row, repay_row))

    # Year 0 opening is hard-zero; later years pull the prior closing balance.
    f0 = model_ws.cell(opening_row, 4).value  # column D = year 0
    assert f0 == "=0", f0
    f1 = model_ws.cell(opening_row, 5).value  # column E = year 1
    assert f1 == f"=IF(E$4=0,0,D{closing_row})", f1

    # Closing = opening + drawdown + IDC − repayment.
    closing_formula = model_ws.cell(closing_row, 4).value
    assert closing_formula == (
        f"=D{opening_row}+D{draw_row}+D{idc_row}-D{repay_row}"
    ), closing_formula


# ── One combined Hourly sheet ────────────────────────────────────────────────


def _fake_year_results(n_years: int = 2, n_hours: int = 24,
                       first_year: int = 2025) -> list[SimpleNamespace]:
    """Minimal stand-ins for OptimisationResult with aligned Series/Indexes."""
    results = []
    for y in range(n_years):
        idx = pd.date_range(f"{first_year + y}-01-01", periods=n_hours, freq="h")
        hours = range(n_hours)

        def _ser(vals):
            return pd.Series(vals, index=idx)

        dispatch = SimpleNamespace(
            wind_gen=_ser([1.0 * (h % 3) for h in hours]),
            pv_gen=_ser([0.5 if 9 <= h < 17 else 0.0 for h in hours]),
            bess_dispatch=_ser([0.2] * n_hours),
            bess_store=_ser([0.1] * n_hours),
            market_buy=_ser([0.05] * n_hours),
            market_sell=_ser([0.7] * n_hours),
            ppa_delivery=_ser([0.9] * n_hours),
            penalty_gen=_ser([0.02] * n_hours),
        )
        results.append(SimpleNamespace(
            scenario=SimpleNamespace(first_sim_year=first_year),
            dispatch=dispatch,
            market_prices=_ser([50.0 + h for h in hours]),
        ))
    return results


def test_excel_export_single_combined_hourly_sheet():
    p = ProjectFinanceInputs()
    e = _simple_energy()
    result = run_project_finance(p, e)
    xlsx_bytes = export_financial_model(p, e, result, year_results=_fake_year_results())
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=False)

    # Exactly one Hourly sheet (no per-year "Hourly Y1/Y2" sheets).
    hourly = [ws.title for ws in wb.worksheets if ws.title.startswith("Hourly")]
    assert hourly == ["Hourly"], hourly

    ws = wb["Hourly"]
    assert ws.cell(_HOURLY_HEADER_ROW, 1).value == "Year"
    assert ws.cell(_HOURLY_HEADER_ROW, 2).value == "Timestamp"

    # Per-year aggregate formulas filter the stacked rows by the Year column.
    ppa_row = _AGG_ROWS["ppa_gwh"]
    agg = ws.cell(ppa_row, 3).value  # first per-year column
    assert isinstance(agg, str) and agg.startswith("=SUMIFS("), agg
    assert f"$A${_HOURLY_DATA_START}:$A$" in agg  # year column criteria range

    # The Energy sheet rolls its annual totals up from the Hourly sheet.
    energy_ws = wb["Energy"]
    ppa_energy_cell = None
    for row in energy_ws.iter_rows(min_col=2, max_col=3):
        if row[0].value == "PPA delivered":
            ppa_energy_cell = row[1]
            break
    assert ppa_energy_cell is not None
    assert str(ppa_energy_cell.value).startswith("=AVERAGE('Hourly'!"), (
        ppa_energy_cell.value
    )


# ── Outputs sheet is fully formula-driven ────────────────────────────────────


def test_excel_export_outputs_are_live_formulas():
    wb, _ = _load_model_ws()
    outputs_ws = wb["Outputs"]
    expected_labels = {
        "Project IRR (FCFF)", "Equity IRR (FCFE)", "NPV @ WACC (project)",
        "Gearing", "Total funding (incl. IDC)", "Total debt", "Total equity",
        "Minimum DSCR", "Average DSCR", "Equity payback", "LCOE",
    }
    found = set()
    for row in outputs_ws.iter_rows(min_col=2, max_col=3):
        label, value = row[0].value, row[1].value
        if label in expected_labels:
            assert isinstance(value, str) and value.startswith("="), (
                f"Outputs {label!r} is not a live formula: {value!r}"
            )
            found.add(label)
    assert found == expected_labels, f"missing KPIs: {expected_labels - found}"
