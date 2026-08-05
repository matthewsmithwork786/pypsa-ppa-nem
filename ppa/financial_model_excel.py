"""Export a streamlined, *live* project-finance workbook.

Produces an ``.xlsx`` that mirrors :mod:`ppa.financial_model`: editable Inputs, a
pre-filled Energy (PyPSA interface) sheet, a single combined Hourly dispatch
sheet, a transposed annual Model sheet and a fully formula-driven Outputs sheet.

Conventions (year 0):
* The Model sheet's periods run **year 0 .. n-1**; year 0 is the base year, so
  no indexation/escalation is applied there (multipliers are 1.0 at year 0 and
  compound ``(1 + rate)^year`` thereafter).
* Devex and capex are broken out **by technology** (wind / solar / BESS) with
  subtotal rows.
* The revenue block is preceded by a *volumes & prices* block: PPA volume,
  merchant volumes and penalty volume, each with an average price per MWh, and
  the revenue rows are formulas multiplying those rows together.
* The debt schedule is shown as a **corkscrew**: opening balance → drawdown →
  IDC/interest → repayment → closing balance, with closing feeding the next
  period's opening.
* Every number on the Outputs sheet is a live formula referencing the Model /
  Energy / Hourly sheets, so editing an input recomputes the KPIs.

The revenue → EBITDA → depreciation → tax → cash-flow chain and the IRR/NPV/DSCR
outputs are written as live Excel formulas. The debt sizing (front-loaded
drawdown, IDC, DSCR tranche split) is circular by nature, so it is written as
toolkit-computed values that the live formulas reference — clearly flagged so it
can be overridden.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ppa.financial_model import (
    ProjectFinanceInputs,
    EnergyInputs,
    ProjectFinanceResult,
    _build_timeline,
)

# ── Styling ──────────────────────────────────────────────────────────────────
_TITLE = Font(bold=True, size=14, color="1F4E78")
_HEADER = Font(bold=True, color="FFFFFF")
_SECTION = Font(bold=True, size=11, color="1F4E78")
_INPUT_FONT = Font(color="0000CC")  # blue = editable input (convention)
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
_PREFILL = PatternFill("solid", fgColor="E2EFDA")
_SECTION_FILL = PatternFill("solid", fgColor="DDEBF7")
_thin = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _pcol(year: int) -> int:
    """Spreadsheet column index for a 0-based model year (year 0 -> col D=4)."""
    return 4 + year


def _text(cell, value):
    """Write *value* into *cell* as literal text, never as a formula.

    openpyxl infers a formula from any string beginning with ``=`` (and would
    emit an unparseable ``<f>`` element that Excel refuses to open). Forcing
    ``data_type="s"`` stores the text verbatim, so a label/note/unit can never
    corrupt the workbook. Returns the cell for chaining.
    """
    cell.value = value
    if isinstance(value, str) and value[:1] in "=+-@":
        cell.data_type = "s"
    return cell


def export_hourly_timeseries(year_results: list) -> bytes:
    """The stacked hourly dispatch as its own workbook.

    25 years x 8760 rows is ~219k rows of thirteen columns: slow to build and
    heavy to move, which made the financial model itself slow to download for no
    benefit to anyone reading the P&L. The hours ship separately so the finance
    workbook stays small, and carry the same per-year aggregate block so they
    remain auditable on their own.
    """
    if not year_results:
        raise ValueError("export_hourly_timeseries needs at least one year of results")

    wb = Workbook()
    wb.remove(wb.active)  # drop openpyxl's default empty sheet
    _write_hourly_sheet(wb, year_results)
    wb.active = 0
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_financial_model(
    p: ProjectFinanceInputs,
    e: EnergyInputs,
    result: ProjectFinanceResult,
    year_results: list | None = None,
) -> bytes:
    """The project-finance workbook.

    Passing `year_results` embeds the hourly sheet and makes the Energy totals
    roll up from it. The app no longer does: the hours ship via
    `export_hourly_timeseries` and the Energy tab carries hard values instead.
    The argument is kept so a single self-contained workbook remains available.
    """
    wb = Workbook()
    inputs_cells = _write_inputs(wb, p)
    # One combined Hourly sheet holding every simulated year stacked vertically;
    # the Energy totals roll up from it when per-year results are available.
    hourly_refs = _write_hourly_sheet(wb, year_results) if year_results else None
    energy_cells = _write_energy(wb, e, hourly_refs)
    model_cells = _write_model(wb, p, e, result, inputs_cells, energy_cells)
    _write_outputs(wb, result, inputs_cells, model_cells)
    _write_notes(wb)

    # Sheet order: headline sheets first, the bulky combined Hourly sheet last.
    _order = {"Outputs": 0, "Inputs": 1, "Energy": 2, "Model": 3, "Notes": 4, "Hourly": 5}
    wb._sheets.sort(key=lambda ws: _order.get(ws.title, 99))
    wb.active = 0

    # Recalculate formulas on open
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Inputs sheet ──────────────────────────────────────────────────────────────


def _write_inputs(wb: Workbook, p: ProjectFinanceInputs) -> dict[str, str]:
    ws = wb.active
    ws.title = "Inputs"
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 60

    _text(ws["B1"], "Financial Model — Inputs")
    ws["B1"].font = _TITLE
    _text(ws["B2"], "Yellow cells are editable assumptions. Costs in A$M/MW (A$M/MWh for BESS).")
    ws["B2"].font = Font(italic=True, color="808080")

    cells: dict[str, str] = {}
    row = 4

    def section(title: str) -> None:
        nonlocal row
        _text(ws.cell(row, 2), title).font = _SECTION
        for c in range(2, 6):
            ws.cell(row, c).fill = _SECTION_FILL
        row += 1

    def field(label: str, key: str, value, unit: str = "", note: str = "") -> None:
        nonlocal row
        _text(ws.cell(row, 2), label)
        vc = ws.cell(row, 3, value)
        vc.fill = _INPUT_FILL
        vc.font = _INPUT_FONT
        vc.border = _BORDER
        if isinstance(value, float):
            vc.number_format = "#,##0.0000" if abs(value) < 10 else "#,##0.00"
        _text(ws.cell(row, 4), unit).font = Font(color="808080")
        if note:
            _text(ws.cell(row, 5), note).font = Font(italic=True, color="A0A0A0")
        cells[key] = f"Inputs!$C${row}"
        row += 1

    section("Build cost")
    field("Onshore wind build cost", "onsw_build_cost", p.onsw_build_cost, "A$M/MW")
    field("Solar PV build cost", "pv_build_cost", p.pv_build_cost, "A$M/MW")
    field("BESS build cost", "bess_build_cost", p.bess_build_cost, "A$M/MWh")
    row += 1
    section("Connection cost")
    field("Onshore wind connection", "onsw_connection_cost", p.onsw_connection_cost, "A$M/MW")
    field("Solar PV connection", "pv_connection_cost", p.pv_connection_cost, "A$M/MW")
    field("BESS connection", "bess_connection_cost", p.bess_connection_cost, "A$M/MWh")
    row += 1
    section("Project development cost (devex)")
    field("Onshore wind devex", "onsw_devex", p.onsw_devex, "A$M/MW")
    field("Solar PV devex", "pv_devex", p.pv_devex, "A$M/MW")
    field("BESS devex", "bess_devex", p.bess_devex, "A$M/MWh")
    row += 1
    section("Fixed O&M (p.a.)")
    field("Onshore wind fixed O&M", "onsw_fixed_om", p.onsw_fixed_om, "A$M/MW")
    field("Solar PV fixed O&M", "pv_fixed_om", p.pv_fixed_om, "A$M/MW")
    field("BESS fixed O&M", "bess_fixed_om", p.bess_fixed_om, "A$M/MWh")
    field("Ancillary services", "ancillary_pct", p.ancillary_pct, "% of revenue")
    row += 1
    section("Timing (years)")
    field("Model duration", "model_duration", p.model_duration, "years")
    field("Development start period", "development_start", p.development_start, "period",
          "FID — devex bullet and construction both start here (1 = year 0)")
    field("Onshore wind construction", "onsw_constr_years", p.onsw_constr_years, "years")
    field("Solar PV construction", "pv_constr_years", p.pv_constr_years, "years")
    field("BESS construction", "bess_constr_years", p.bess_constr_years, "years")
    field("Operating life", "operating_life", p.operating_life, "years")
    row += 1
    section("Revenue")
    field("PPA contract tenor", "ppa_tenor", p.ppa_tenor, "years")
    field("PPA tariff (base)", "ppa_tariff", p.ppa_tariff, "A$/MWh")
    field("Penalty multiple", "penalty_multiple", p.penalty_multiple, "×")
    field("LGC / GO price", "lgc_price", p.lgc_price, "A$/MWh")
    row += 1
    section("Indexation (% p.a.) — applied from year 1; year 0 is the base year")
    field("Cost inflation", "cost_inflation", p.cost_inflation, "%")
    field("PPA & LGC indexation", "ppa_indexation", p.ppa_indexation, "%")
    field("Solar-hour price inflation", "solar_price_inflation", p.solar_price_inflation, "%")
    field("Non-solar-hour price inflation", "nonsolar_price_inflation", p.nonsolar_price_inflation, "%")
    row += 1
    section("Project finance")
    field("Debt repayment tenor", "debt_tenor", p.debt_tenor, "years")
    field("Debt rate", "debt_rate", p.debt_rate, "%")
    field("DSCR hurdle (contracted)", "dscr_contracted", p.dscr_contracted, "ratio")
    field("DSCR hurdle (uncontracted)", "dscr_uncontracted", p.dscr_uncontracted, "ratio")
    field("Max gearing (contracted)", "max_gearing_contracted", p.max_gearing_contracted, "%")
    field("Max gearing (uncontracted)", "max_gearing_uncontracted", p.max_gearing_uncontracted, "%")
    row += 1
    section("Depreciation & tax")
    field("Book depreciation rate", "book_depreciation_rate", p.book_depreciation_rate, "%")
    field("Tax depreciation rate", "tax_depreciation_rate", p.tax_depreciation_rate, "%")
    field("Corporate tax rate", "corp_tax_rate", p.corp_tax_rate, "%")
    field("Discount rate (WACC)", "discount_rate", p.discount_rate, "%")

    return cells


# ── Combined Hourly dispatch sheet ────────────────────────────────────────────

# Fixed row layout for the per-year annual aggregate block, so the Energy sheet
# can reference these cells by address. The block spans one column per year.
_AGG_HEADER_ROW = 3
_AGG_ROWS = {
    "scale": 4,
    "ppa_gwh": 5,
    "excess_solar_gwh": 6,
    "excess_nonsolar_gwh": 7,
    "penalty_gwh": 8,
    "total_solar_gwh": 9,
    "total_nonsolar_gwh": 10,
    "sell_solar_price": 11,
    "sell_nonsolar_price": 12,
    "purchase_price": 13,
    "marketbuy_gwh": 14,
}
_HOURLY_HEADER_ROW = 16
_HOURLY_DATA_START = 17

# Hourly data columns (1-based): year, timestamp, hour, then the energy/price
# series. The year column drives the SUMIFS-based per-year aggregates above.
_HOURLY_COLS = [
    "Year", "Timestamp", "Hour", "Wind (MWh)", "PV (MWh)", "BESS discharge (MWh)",
    "BESS charge (MWh)", "Total generation (MWh)", "Market buy (MWh)",
    "Market sell (MWh)", "PPA delivered (MWh)", "Penalty (MWh)", "Price (A$/MWh)",
]
_C_YEAR, _C_HOUR = "A", "C"
_C_TOTAL, _C_BUY, _C_SELL = "H", "I", "J"
_C_PPA, _C_PEN, _C_PRICE = "K", "L", "M"


def _year_column(year_idx_1based: int) -> str:
    """Column letter of the per-year aggregate block for a 1-based year index
    (year 1 -> column C = 3)."""
    return get_column_letter(2 + year_idx_1based)


def _write_hourly_sheet(wb: Workbook, year_results: list) -> dict[str, list[str]]:
    """Write ONE combined hourly-dispatch sheet and return, per metric, the list
    of per-year aggregate cell references (for the Energy sheet to average).

    All years' hourly rows are stacked vertically in the same columns, headed by
    a ``Year`` column. A top-of-sheet aggregate block spreads one column per year
    and every aggregate is a live ``SUMIFS``/``SUMPRODUCT`` that filters the
    stacked rows by year — so the totals feeding the financial model are
    auditable roll-ups of the hourly data, computed live by Excel."""
    import numpy as np  # noqa: F401  (kept for parity / future use)

    refs: dict[str, list[str]] = {k: [] for k in _AGG_ROWS if k != "scale"}

    first_year = getattr(year_results[0].scenario, "first_sim_year", 0)
    year_labels = []
    n_rows = []
    for idx, res in enumerate(year_results, start=1):
        year_labels.append((first_year + idx - 1) if first_year else idx)
        n_rows.append(len(res.dispatch.wind_gen))

    n_years = len(year_results)
    last = _HOURLY_DATA_START + sum(n_rows) - 1
    ws = wb.create_sheet("Hourly")
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 18

    ws["A1"] = "Hourly dispatch — all simulated years (stacked)"
    ws["A1"].font = _TITLE
    _text(ws["A2"],
          "Each year's hourly rows are stacked with a Year column. The annual "
          "aggregates above filter by year (SUMIFS) and scale by 8760/hours.").font = (
        Font(italic=True, color="808080"))

    # ── Aggregate block header: one column per year ──────────────────────────
    _text(ws.cell(_AGG_HEADER_ROW, 1), "Metric").font = _HEADER
    ws.cell(_AGG_HEADER_ROW, 1).fill = _HEADER_FILL
    _text(ws.cell(_AGG_HEADER_ROW, 2), "Unit").font = _HEADER
    ws.cell(_AGG_HEADER_ROW, 2).fill = _HEADER_FILL
    for j, yl in enumerate(year_labels, start=1):
        c = ws.cell(_AGG_HEADER_ROW, 2 + j, yl)
        c.font = _HEADER
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    # ── Aggregate labels / units ─────────────────────────────────────────────
    _AGG_LABELS = {
        "scale": "Annualisation factor (8760 / hours)",
        "ppa_gwh": "PPA delivered",
        "excess_solar_gwh": "Excess sold — solar hours",
        "excess_nonsolar_gwh": "Excess sold — non-solar hours",
        "penalty_gwh": "Penalty (undelivered)",
        "total_solar_gwh": "Total generation — solar hours",
        "total_nonsolar_gwh": "Total generation — non-solar hours",
        "sell_solar_price": "Merchant capture — solar hours",
        "sell_nonsolar_price": "Merchant capture — non-solar hours",
        "purchase_price": "Market purchase price",
        "marketbuy_gwh": "Market purchase volume",
    }
    _AGG_UNITS = {
        "scale": "×",
        "ppa_gwh": "GWh p.a.",
        "excess_solar_gwh": "GWh p.a.",
        "excess_nonsolar_gwh": "GWh p.a.",
        "penalty_gwh": "GWh p.a.",
        "total_solar_gwh": "GWh p.a.",
        "total_nonsolar_gwh": "GWh p.a.",
        "sell_solar_price": "A$/MWh",
        "sell_nonsolar_price": "A$/MWh",
        "purchase_price": "A$/MWh",
        "marketbuy_gwh": "GWh p.a.",
    }
    for key, label in _AGG_LABELS.items():
        r = _AGG_ROWS[key]
        _text(ws.cell(r, 1), label).font = Font(bold=True)
        _text(ws.cell(r, 2), _AGG_UNITS[key]).font = Font(color="808080", size=9)

    # ── Per-year aggregate formulas ──────────────────────────────────────────
    year_rng = f"${_C_YEAR}${_HOURLY_DATA_START}:${_C_YEAR}${last}"
    hour_rng = f"${_C_HOUR}${_HOURLY_DATA_START}:${_C_HOUR}${last}"
    rng = lambda col: f"${col}${_HOURLY_DATA_START}:${col}${last}"

    solar_w = f"(({hour_rng}>=9)*({hour_rng}<17))"
    nonsolar_w = f"((({hour_rng}<9)+({hour_rng}>=17)))"
    sell_rng = rng(_C_SELL)
    total_rng = rng(_C_TOTAL)
    buy_rng = rng(_C_BUY)
    price_rng = rng(_C_PRICE)
    ppa_rng = rng(_C_PPA)
    pen_rng = rng(_C_PEN)

    for j in range(1, n_years + 1):
        cl = _year_column(j)
        crit = f"{cl}${_AGG_HEADER_ROW}"
        scale_cell = f"{cl}${_AGG_ROWS['scale']}"
        yr_mask = f"({year_rng}={crit})"

        def aggcell(row: int, formula: str, fmt: str = "#,##0.00") -> None:
            c = ws.cell(row, 2 + j, formula)
            c.number_format = fmt
            c.fill = _PREFILL

        aggcell(_AGG_ROWS["scale"],
                f"=8760/COUNTIF({year_rng},{crit})", "0.0000")
        aggcell(_AGG_ROWS["ppa_gwh"],
                f"=SUMIFS({ppa_rng},{year_rng},{crit})*{scale_cell}/1000")
        solar_sell = f'SUMIFS({sell_rng},{year_rng},{crit},{hour_rng},">=9",{hour_rng},"<17")'
        aggcell(_AGG_ROWS["excess_solar_gwh"],
                f"={solar_sell}*{scale_cell}/1000")
        aggcell(_AGG_ROWS["excess_nonsolar_gwh"],
                f"=(SUMIFS({sell_rng},{year_rng},{crit})*{scale_cell}/1000)"
                f"-{cl}${_AGG_ROWS['excess_solar_gwh']}")
        aggcell(_AGG_ROWS["penalty_gwh"],
                f"=SUMIFS({pen_rng},{year_rng},{crit})*{scale_cell}/1000")
        solar_total = f'SUMIFS({total_rng},{year_rng},{crit},{hour_rng},">=9",{hour_rng},"<17")'
        aggcell(_AGG_ROWS["total_solar_gwh"],
                f"={solar_total}*{scale_cell}/1000")
        aggcell(_AGG_ROWS["total_nonsolar_gwh"],
                f"=(SUMIFS({total_rng},{year_rng},{crit})*{scale_cell}/1000)"
                f"-{cl}${_AGG_ROWS['total_solar_gwh']}")
        # Volume-weighted capture prices (solar / non-solar hours), guarded for /0
        aggcell(_AGG_ROWS["sell_solar_price"],
                f"=IFERROR(SUMPRODUCT({solar_w}*{yr_mask}*{sell_rng}*{price_rng})"
                f"/SUMPRODUCT({solar_w}*{yr_mask}*{sell_rng}),0)", "0.00")
        aggcell(_AGG_ROWS["sell_nonsolar_price"],
                f"=IFERROR(SUMPRODUCT({nonsolar_w}*{yr_mask}*{sell_rng}*{price_rng})"
                f"/SUMPRODUCT({nonsolar_w}*{yr_mask}*{sell_rng}),0)", "0.00")
        aggcell(_AGG_ROWS["purchase_price"],
                f"=IFERROR(SUMPRODUCT({yr_mask}*{buy_rng}*{price_rng})"
                f"/SUMPRODUCT({yr_mask}*{buy_rng}),0)", "0.00")
        aggcell(_AGG_ROWS["marketbuy_gwh"],
                f"=SUMIFS({buy_rng},{year_rng},{crit})*{scale_cell}/1000")

        for k in refs:
            refs[k].append(f"'Hourly'!{cl}${_AGG_ROWS[k]}")

    # ── Stacked hourly data ───────────────────────────────────────────────────
    for j, name in enumerate(_HOURLY_COLS, start=1):
        cell = ws.cell(_HOURLY_HEADER_ROW, j, name)
        cell.font = _HEADER
        cell.fill = _HEADER_FILL

    cur = _HOURLY_DATA_START
    for idx, res in enumerate(year_results, start=1):
        d = res.dispatch
        prices = res.market_prices
        wind = d.wind_gen.to_numpy()
        pv = d.pv_gen.to_numpy()
        bess_dis = d.bess_dispatch.to_numpy()
        bess_chg = d.bess_store.to_numpy()
        buy = d.market_buy.to_numpy()
        sell = d.market_sell.to_numpy()
        ppa = d.ppa_delivery.to_numpy()
        pen = d.penalty_gen.to_numpy()
        price = prices.to_numpy()
        total = wind + pv + bess_dis
        index = d.wind_gen.index
        hours = index.hour
        year_val = year_labels[idx - 1]
        for i in range(len(wind)):
            ws.cell(cur, 1, year_val)
            ws.cell(cur, 2, index[i].strftime("%Y-%m-%d %H:%M"))
            ws.cell(cur, 3, int(hours[i]))
            ws.cell(cur, 4, round(float(wind[i]), 3))
            ws.cell(cur, 5, round(float(pv[i]), 3))
            ws.cell(cur, 6, round(float(bess_dis[i]), 3))
            ws.cell(cur, 7, round(float(bess_chg[i]), 3))
            ws.cell(cur, 8, round(float(total[i]), 3))
            ws.cell(cur, 9, round(float(buy[i]), 3))
            ws.cell(cur, 10, round(float(sell[i]), 3))
            ws.cell(cur, 11, round(float(ppa[i]), 3))
            ws.cell(cur, 12, round(float(pen[i]), 3))
            ws.cell(cur, 13, round(float(price[i]), 3))
            cur += 1
    ws.freeze_panes = f"A{_HOURLY_DATA_START}"

    return refs


# ── Energy sheet (PyPSA interface) ───────────────────────────────────────────


def _write_energy(
    wb: Workbook,
    e: EnergyInputs,
    hourly_refs: dict[str, list[str]] | None = None,
) -> dict[str, str]:
    ws = wb.create_sheet("Energy")
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 10

    _text(ws["B1"], "PyPSA Energy Model Results")
    ws["B1"].font = _TITLE
    _text(ws["B2"],
          f"Scenario: {e.name}. "
          + ("Annual totals are the average of the per-year sums on the Hourly sheet."
             if hourly_refs else
             "Annual totals are hard values: the average across the simulated "
             "years of each year's hourly dispatch. They are the same numbers "
             "the separate hourly-timeseries workbook sums to, and everything "
             "on Model and Outputs is driven from them."))
    ws["B2"].font = Font(italic=True, color="808080")

    cells: dict[str, str] = {}
    row = 4

    def field(label: str, key: str, value, unit: str = "") -> None:
        nonlocal row
        _text(ws.cell(row, 2), label)
        # Summable metrics roll up from the Hourly sheet when available.
        if hourly_refs and key in hourly_refs:
            vc = ws.cell(row, 3, f"=AVERAGE({','.join(hourly_refs[key])})")
        else:
            vc = ws.cell(row, 3, value)
        vc.fill = _PREFILL
        vc.border = _BORDER
        if isinstance(value, float):
            vc.number_format = "#,##0.00"
        _text(ws.cell(row, 4), unit).font = Font(color="808080")
        cells[key] = f"Energy!$C${row}"
        row += 1

    field("Onshore wind capacity", "onsw_mw", e.onsw_mw, "MW")
    field("Solar PV capacity", "pv_mw", e.pv_mw, "MW")
    field("BESS power", "bess_mw", e.bess_mw, "MW")
    field("BESS energy", "bess_mwh", e.bess_mwh, "MWh")
    field("Offtaker load", "load_mw", e.load_mw, "MW")
    row += 1
    field("PPA delivered", "ppa_gwh", e.ppa_gwh, "GWh p.a.")
    field("Excess sold — solar hours", "excess_solar_gwh", e.excess_solar_gwh, "GWh p.a.")
    field("Excess sold — non-solar hours", "excess_nonsolar_gwh", e.excess_nonsolar_gwh, "GWh p.a.")
    field("Penalty (undelivered)", "penalty_gwh", e.penalty_gwh, "GWh p.a.")
    row += 1
    field("Total generation — solar hours", "total_solar_gwh", e.total_solar_gwh, "GWh p.a.")
    field("Total generation — non-solar hours", "total_nonsolar_gwh", e.total_nonsolar_gwh, "GWh p.a.")
    row += 1
    field("Merchant capture — solar hours", "sell_solar_price", e.sell_solar_price, "A$/MWh")
    field("Merchant capture — non-solar hours", "sell_nonsolar_price", e.sell_nonsolar_price, "A$/MWh")
    field("Market purchase price", "purchase_price", e.purchase_price, "A$/MWh")
    field("Market purchase volume", "marketbuy_gwh", e.marketbuy_gwh, "GWh p.a.")

    return cells


# ── Model sheet (transposed; live formulas) ──────────────────────────────────


def _write_model(
    wb: Workbook,
    p: ProjectFinanceInputs,
    e: EnergyInputs,
    result: ProjectFinanceResult,
    I: dict[str, str],
    E: dict[str, str],
) -> dict[str, str]:
    ws = wb.create_sheet("Model")
    n = p.model_duration
    tl = _build_timeline(p)
    sc = result.schedule

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 8

    _text(ws["B1"], "Annual Project-Finance Model")
    ws["B1"].font = _TITLE
    _text(ws["B2"],
          "Years run 0..n-1; year 0 is the base year (no indexation). Revenue, "
          "volumes, opex, depreciation, tax and cash flows are live formulas. "
          "Debt drawdown/IDC and tranche split are toolkit-sized values (green).")
    ws["B2"].font = Font(italic=True, color="808080")

    # Period header — year 0 .. n-1
    hdr = 4
    _text(ws.cell(hdr, 2), "Year").font = _HEADER
    ws.cell(hdr, 2).fill = _HEADER_FILL
    _text(ws.cell(hdr, 3), "Unit").font = _HEADER
    ws.cell(hdr, 3).fill = _HEADER_FILL
    for year in range(0, n):
        c = ws.cell(hdr, _pcol(year), year)
        c.font = _HEADER
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    R: dict[str, int] = {}
    row = hdr + 1

    def col(year: int) -> str:
        return get_column_letter(_pcol(year))

    def label_row(name: str, label: str, unit: str = "", section: bool = False) -> int:
        nonlocal row
        r = row
        cell = _text(ws.cell(r, 2), label)
        if section:
            cell.font = _SECTION
            for cc in range(2, _pcol(n - 1) + 1):
                ws.cell(r, cc).fill = _SECTION_FILL
        _text(ws.cell(r, 3), unit).font = Font(color="808080", size=9)
        R[name] = r
        row += 1
        return r

    def put_formula(name: str, fn, fmt: str = "#,##0.0", value_fill: bool = False) -> None:
        r = R[name]
        for year in range(0, n):
            cell = ws.cell(r, _pcol(year), fn(year, col(year)))
            cell.number_format = fmt
            if value_fill:
                cell.fill = _PREFILL

    def put_values(name: str, arr, fmt: str = "#,##0.0", value_fill: bool = True) -> None:
        r = R[name]
        for year in range(0, n):
            cell = ws.cell(r, _pcol(year), round(float(arr[year]), 6))
            cell.number_format = fmt
            if value_fill:
                cell.fill = _PREFILL

    # ── Flags (values) ───────────────────────────────────────────────────────
    label_row("flags", "Flags", section=True)
    label_row("ops_flag", "Operations flag", "0/1")
    put_values("ops_flag", sc["ops_flag"], "0")
    label_row("ppa_flag", "PPA flag", "0/1")
    put_values("ppa_flag", sc["ppa_flag"], "0")
    nonppa = sc["ops_flag"] - sc["ppa_flag"]
    label_row("nonppa_flag", "Post-PPA flag", "0/1")
    put_values("nonppa_flag", nonppa, "0")
    debt_flag = ((result.periods >= tl.ops_start) & (result.periods <= tl.debt_end)).astype(float)
    label_row("debt_flag", "Debt repayment flag", "0/1")
    put_values("debt_flag", debt_flag, "0")

    # ── Indexation (formulas; year 0 → multiplier 1.0) ───────────────────────
    label_row("index", "Indexation multiples", section=True)
    label_row("cost_idx", "Cost inflation", "×")
    put_formula("cost_idx", lambda pr, cl: f"=(1+{I['cost_inflation']})^{cl}${hdr}", "0.000")
    label_row("ppa_idx", "PPA & LGC", "×")
    put_formula("ppa_idx", lambda pr, cl: f"=(1+{I['ppa_indexation']})^{cl}${hdr}", "0.000")
    label_row("solar_idx", "Solar-hour price", "×")
    put_formula("solar_idx", lambda pr, cl: f"=(1+{I['solar_price_inflation']})^{cl}${hdr}", "0.000")
    label_row("nonsolar_idx", "Non-solar-hour price", "×")
    put_formula("nonsolar_idx", lambda pr, cl: f"=(1+{I['nonsolar_price_inflation']})^{cl}${hdr}", "0.000")

    # ── Capital spend, by technology (live: cost inputs × capacity × indexation)
    def _fracs(first: int, last: int) -> list[float]:
        arr = [0.0] * n
        if last >= first:
            per = 1.0 / (last - first + 1)
            for yy in range(first, last + 1):
                arr[yy] = per
        return arr

    onsw_con_f = _fracs(*tl.tech_constr(p.onsw_constr_years))
    pv_con_f = _fracs(*tl.tech_constr(p.pv_constr_years))
    bess_con_f = _fracs(*tl.tech_constr(p.bess_constr_years))

    def _devex_tech_fn(cost: str, cap: str):
        # Devex is a single bullet at FID. The IF keys off the live
        # `development_start` input (1-based) mapped to a 0-based year, so
        # editing it in Excel moves the bullet.
        return (
            lambda pr, cl: (
                f"=IF({cl}${hdr}={I['development_start']}-1,"
                f"{cl}{R['cost_idx']}*({I[cost]}*{E[cap]}),0)"
            )
        )

    def _capex_tech_fn(frac: list[float], build: str, conn: str, cap: str):
        return (
            lambda pr, cl: (
                f"={cl}{R['cost_idx']}*("
                f"{frac[pr]}*({I[build]}+{I[conn]})*{E[cap]})"
            )
        )

    label_row("devex_sec", "Capital spend — development (devex)", "A$M", section=True)
    label_row("devex_onsw", "Devex — Onshore wind", "A$M")
    put_formula("devex_onsw", _devex_tech_fn("onsw_devex", "onsw_mw"))
    label_row("devex_pv", "Devex — Solar PV", "A$M")
    put_formula("devex_pv", _devex_tech_fn("pv_devex", "pv_mw"))
    label_row("devex_bess", "Devex — BESS", "A$M")
    put_formula("devex_bess", _devex_tech_fn("bess_devex", "bess_mwh"))
    label_row("devex", "Devex", "A$M")
    put_formula("devex", lambda pr, cl: (
        f"={cl}{R['devex_onsw']}+{cl}{R['devex_pv']}+{cl}{R['devex_bess']}"
    ))

    label_row("capex_sec", "Capital spend — construction (capex)", "A$M", section=True)
    label_row("capex_onsw", "Capex — Onshore wind", "A$M")
    put_formula("capex_onsw", _capex_tech_fn(onsw_con_f, "onsw_build_cost", "onsw_connection_cost", "onsw_mw"))
    label_row("capex_pv", "Capex — Solar PV", "A$M")
    put_formula("capex_pv", _capex_tech_fn(pv_con_f, "pv_build_cost", "pv_connection_cost", "pv_mw"))
    label_row("capex_bess", "Capex — BESS", "A$M")
    put_formula("capex_bess", _capex_tech_fn(bess_con_f, "bess_build_cost", "bess_connection_cost", "bess_mwh"))
    label_row("capex", "Capex", "A$M")
    put_formula("capex", lambda pr, cl: (
        f"={cl}{R['capex_onsw']}+{cl}{R['capex_pv']}+{cl}{R['capex_bess']}"
    ))

    label_row("capital_spend", "Total capital spend", "A$M")
    put_formula("capital_spend", lambda pr, cl: f"={cl}{R['devex']}+{cl}{R['capex']}")

    # ── Volumes & prices (shown first; revenue multiplies these rows) ────────
    label_row("vols", "Volumes & prices", section=True)
    label_row("ppa_vol", "PPA delivered volume", "GWh")
    put_formula("ppa_vol", lambda pr, cl: f"={cl}{R['ppa_flag']}*{E['ppa_gwh']}")
    label_row("merch_solar_vol", "Merchant volume — solar hours", "GWh")
    put_formula("merch_solar_vol", lambda pr, cl: (
        f"={cl}{R['ppa_flag']}*{E['excess_solar_gwh']}+{cl}{R['nonppa_flag']}*{E['total_solar_gwh']}"
    ))
    label_row("merch_nonsolar_vol", "Merchant volume — non-solar hours", "GWh")
    put_formula("merch_nonsolar_vol", lambda pr, cl: (
        f"={cl}{R['ppa_flag']}*{E['excess_nonsolar_gwh']}+{cl}{R['nonppa_flag']}*{E['total_nonsolar_gwh']}"
    ))
    label_row("merch_vol", "Merchant volume — total", "GWh")
    put_formula("merch_vol", lambda pr, cl: f"={cl}{R['merch_solar_vol']}+{cl}{R['merch_nonsolar_vol']}")
    label_row("pen_vol", "Penalty generation volume", "GWh")
    put_formula("pen_vol", lambda pr, cl: f"={cl}{R['ppa_flag']}*{E['penalty_gwh']}")
    row += 1
    label_row("ppa_price", "Average PPA price", "A$/MWh")
    put_formula("ppa_price", lambda pr, cl: f"={I['ppa_tariff']}*{cl}{R['ppa_idx']}")
    esc = p.escalate_merchant_prices
    label_row("merch_solar_price", "Average merchant price — solar hours", "A$/MWh")
    put_formula("merch_solar_price",
                lambda pr, cl: f"={E['sell_solar_price']}*{cl}{R['solar_idx']}" if esc else f"={E['sell_solar_price']}")
    label_row("merch_nonsolar_price", "Average merchant price — non-solar hours", "A$/MWh")
    put_formula("merch_nonsolar_price",
                lambda pr, cl: f"={E['sell_nonsolar_price']}*{cl}{R['nonsolar_idx']}" if esc else f"={E['sell_nonsolar_price']}")
    label_row("pen_price", "Average penalty price", "A$/MWh")
    put_formula("pen_price", lambda pr, cl: f"={I['ppa_tariff']}*{I['penalty_multiple']}*{cl}{R['ppa_idx']}")
    label_row("lgc_price", "LGC / GO price", "A$/MWh")
    put_formula("lgc_price", lambda pr, cl: f"={I['lgc_price']}*{cl}{R['ppa_idx']}")

    # ── Revenue (live formulas over the volume/price rows) ───────────────────
    label_row("revenue", "Revenue", "A$M", section=True)
    label_row("ppa_rev", "PPA revenue", "A$M")
    put_formula("ppa_rev", lambda pr, cl: (
        f"={cl}{R['ppa_vol']}*1000*{cl}{R['ppa_price']}/1000000"
    ))
    label_row("merch_solar_rev", "Merchant revenue — solar hours", "A$M")
    put_formula("merch_solar_rev", lambda pr, cl: (
        f"={cl}{R['merch_solar_vol']}*1000*{cl}{R['merch_solar_price']}/1000000"
    ))
    label_row("merch_nonsolar_rev", "Merchant revenue — non-solar hours", "A$M")
    put_formula("merch_nonsolar_rev", lambda pr, cl: (
        f"={cl}{R['merch_nonsolar_vol']}*1000*{cl}{R['merch_nonsolar_price']}/1000000"
    ))
    label_row("merch_rev", "Merchant revenue — subtotal", "A$M")
    put_formula("merch_rev", lambda pr, cl: f"={cl}{R['merch_solar_rev']}+{cl}{R['merch_nonsolar_rev']}")
    label_row("lgc_rev", "LGC / GO revenue", "A$M")
    put_formula("lgc_rev", lambda pr, cl: (
        f"={cl}{R['merch_vol']}*1000*{cl}{R['lgc_price']}/1000000"
    ))
    label_row("penalty_cost", "Penalty cost", "A$M")
    put_formula("penalty_cost", lambda pr, cl: (
        f"={cl}{R['pen_vol']}*1000*{cl}{R['pen_price']}/1000000"
    ))
    label_row("net_contracted", "Net contracted revenue", "A$M")
    put_formula("net_contracted", lambda pr, cl: f"={cl}{R['ppa_rev']}-{cl}{R['penalty_cost']}")
    label_row("net_uncontracted", "Net uncontracted revenue", "A$M")
    put_formula("net_uncontracted", lambda pr, cl: f"={cl}{R['merch_rev']}+{cl}{R['lgc_rev']}")
    label_row("total_rev", "Total revenue", "A$M")
    put_formula("total_rev", lambda pr, cl: f"={cl}{R['net_contracted']}+{cl}{R['net_uncontracted']}")

    # ── Opex / EBITDA (formulas) ─────────────────────────────────────────────
    label_row("opex_sec", "Operating costs", "A$M", section=True)
    fixed_om_expr = (
        f"({I['onsw_fixed_om']}*{E['onsw_mw']}+{I['pv_fixed_om']}*{E['pv_mw']}+{I['bess_fixed_om']}*{E['bess_mwh']})"
    )
    label_row("fixed_om", "Fixed O&M", "A$M")
    put_formula("fixed_om", lambda pr, cl: f"={cl}{R['ops_flag']}*{fixed_om_expr}")
    label_row("ancillary", "Ancillary services", "A$M")
    put_formula("ancillary", lambda pr, cl: f"={I['ancillary_pct']}*{cl}{R['total_rev']}")
    label_row("opex", "Total O&M expenses", "A$M")
    put_formula("opex", lambda pr, cl: f"={cl}{R['fixed_om']}+{cl}{R['ancillary']}")
    label_row("ebitda", "EBITDA", "A$M")
    put_formula("ebitda", lambda pr, cl: f"={cl}{R['total_rev']}-{cl}{R['opex']}")

    # ── Debt corkscrew (opening → drawdown/IDC/interest/repayment → closing) ─
    label_row("debt", "Debt schedule (corkscrew)", "A$M", section=True)
    for _name in ("debt_opening", "debt_draw", "debt_idc", "debt_interest",
                  "debt_repay", "debt_closing"):
        label_row(_name, "")
    # fill in labels after rows are known (closing row is referenced by opening)
    for _name, _label in (
        ("debt_opening", "Opening balance"),
        ("debt_draw", "Drawdown (construction)"),
        ("debt_idc", "Interest during construction"),
        ("debt_interest", "Term loan interest"),
        ("debt_repay", "Repayment (principal)"),
        ("debt_closing", "Closing balance"),
    ):
        _text(ws.cell(R[_name], 2), _label)
    put_formula("debt_opening", lambda pr, cl: (
        "=0" if pr == 0 else f"=IF({cl}${hdr}=0,0,{col(pr-1)}{R['debt_closing']})"
    ))
    put_values("debt_draw", sc["debt_draw"])
    put_values("debt_idc", sc["idc"])
    put_values("debt_interest", sc["interest"])
    put_values("debt_repay", sc["loan_repay"])
    put_formula("debt_closing", lambda pr, cl: (
        f"={cl}{R['debt_opening']}+{cl}{R['debt_draw']}+{cl}{R['debt_idc']}-{cl}{R['debt_repay']}"
    ))

    # ── Depreciation (live, straight-line capped at the asset base) ───────────
    firstcol, lastcol = col(0), col(n - 1)
    label_row("dep", "Depreciation", "A$M", section=True)

    # Asset bases (live): tax = capex only; book = devex + capex + capitalised IDC.
    label_row("tax_base", "Tax asset base", "A$M")
    ws.cell(R["tax_base"], _pcol(0),
            f"=SUM({firstcol}{R['capex']}:{lastcol}{R['capex']})").number_format = "#,##0.0"
    label_row("book_base", "Book asset base", "A$M")
    ws.cell(R["book_base"], _pcol(0), (
        f"=SUM({firstcol}{R['devex']}:{lastcol}{R['devex']})"
        f"+SUM({firstcol}{R['capex']}:{lastcol}{R['capex']})"
        f"+SUM({firstcol}{R['debt_idc']}:{lastcol}{R['debt_idc']})"
    )).number_format = "#,##0.0"

    def _dep_fn(self_row: int, base_row: int, rate_cell: str):
        # Straight-line at `rate` on the asset base, but never depreciate more
        # than the remaining book value (cumulative prior depreciation in-row).
        base = f"${firstcol}${base_row}"

        def fn(pr: int, cl: str) -> str:
            prior = "0" if pr == 0 else f"SUM(${firstcol}{self_row}:{col(pr-1)}{self_row})"
            return (
                f"={cl}{R['ops_flag']}*MIN({base}*{rate_cell},MAX({base}-{prior},0))"
            )
        return fn

    label_row("tax_dep", "Tax depreciation", "A$M")
    put_formula("tax_dep", _dep_fn(R["tax_dep"], R["tax_base"], I["tax_depreciation_rate"]))
    label_row("book_dep", "Book depreciation", "A$M")
    put_formula("book_dep", _dep_fn(R["book_dep"], R["book_base"], I["book_depreciation_rate"]))

    # ── P&L tax (live, with loss carry-forward) ───────────────────────────────
    label_row("pl", "Profit & loss", "A$M", section=True)
    label_row("ebit", "EBIT", "A$M")
    put_formula("ebit", lambda pr, cl: f"={cl}{R['ebitda']}-{cl}{R['book_dep']}")
    label_row("pbt", "Profit before tax", "A$M")
    put_formula("pbt", lambda pr, cl: f"={cl}{R['ebitda']}-{cl}{R['debt_interest']}-{cl}{R['book_dep']}")
    label_row("taxable", "Taxable income", "A$M")
    put_formula("taxable", lambda pr, cl: f"={cl}{R['ebitda']}-{cl}{R['debt_interest']}-{cl}{R['tax_dep']}")
    label_row("carry", "Carry-forward losses", "A$M")
    put_formula("carry", lambda pr, cl: (
        f"=MIN(0,{cl}{R['taxable']})" if pr == 0
        else f"=MIN(0,{cl}{R['taxable']}+{col(pr-1)}{R['carry']})"
    ))
    label_row("tax", "Income tax", "A$M")
    put_formula("tax", lambda pr, cl: (
        f"=MAX(0,{cl}{R['taxable']})*{I['corp_tax_rate']}" if pr == 0
        else f"=MAX(0,{cl}{R['taxable']}+{col(pr-1)}{R['carry']})*{I['corp_tax_rate']}"
    ))
    label_row("pat", "Profit after tax", "A$M")
    put_formula("pat", lambda pr, cl: f"={cl}{R['pbt']}-{cl}{R['tax']}")
    label_row("net_margin", "Net margin", "%")
    put_formula("net_margin", lambda pr, cl: (
        f"=IF({cl}{R['total_rev']}>0,{cl}{R['pat']}/{cl}{R['total_rev']},\"\")"
    ), "0.0%")

    # ── Balance sheet (simplified; no cash or working capital) ───────────────
    label_row("bs", "Balance sheet", "A$M", section=True)
    label_row("gross_fa", "Fixed assets (gross, incl. capitalised IDC)", "A$M")
    put_formula("gross_fa", lambda pr, cl: (
        f"=SUM(${firstcol}{R['capital_spend']}:{cl}{R['capital_spend']})"
        f"+SUM(${firstcol}{R['debt_idc']}:{cl}{R['debt_idc']})"
    ))
    label_row("accum_dep", "Accumulated depreciation", "A$M")
    put_formula("accum_dep", lambda pr, cl: f"=SUM(${firstcol}{R['book_dep']}:{cl}{R['book_dep']})")
    label_row("net_fa", "Net fixed assets", "A$M")
    put_formula("net_fa", lambda pr, cl: f"={cl}{R['gross_fa']}-{cl}{R['accum_dep']}")
    label_row("net_debt", "Net debt", "A$M")
    put_formula("net_debt", lambda pr, cl: f"={cl}{R['debt_closing']}")
    label_row("net_equity", "Shareholders' funds", "A$M")
    put_formula("net_equity", lambda pr, cl: f"={cl}{R['net_fa']}-{cl}{R['net_debt']}")

    # ── Cash-flow statement ───────────────────────────────────────────────────
    label_row("cf", "Cash flow statement", "A$M", section=True)
    label_row("ocf", "Cash flow from operations", "A$M")
    put_formula("ocf", lambda pr, cl: f"={cl}{R['ops_flag']}*({cl}{R['ebitda']}-{cl}{R['tax']})")
    label_row("investing_cf", "Cash flow from investing", "A$M")
    put_formula("investing_cf", lambda pr, cl: f"=-{cl}{R['capital_spend']}")
    label_row("equity_spend", "Equity investment", "A$M")
    put_formula("equity_spend", lambda pr, cl: f"={cl}{R['capital_spend']}-{cl}{R['debt_draw']}")
    label_row("financing_cf", "Cash flow from financing", "A$M")
    put_formula("financing_cf", lambda pr, cl: (
        f"={cl}{R['debt_draw']}-{cl}{R['debt_interest']}-{cl}{R['debt_repay']}+{cl}{R['equity_spend']}"
    ))
    label_row("net_cashflow", "Net cash flow", "A$M")
    put_formula("net_cashflow", lambda pr, cl: f"={cl}{R['ocf']}+{cl}{R['investing_cf']}+{cl}{R['financing_cf']}")
    label_row("cum_net_cf", "Cumulative net cash flow", "A$M")
    put_formula("cum_net_cf", lambda pr, cl: f"=SUM(${firstcol}{R['net_cashflow']}:{cl}{R['net_cashflow']})")

    # ── Returns (formulas) ────────────────────────────────────────────────────
    label_row("returns", "Returns", "A$M", section=True)
    label_row("fcff", "FCFF (project)", "A$M")
    put_formula("fcff", lambda pr, cl: (
        f"={cl}{R['ops_flag']}*({cl}{R['ebitda']}-{cl}{R['tax']})-{cl}{R['capital_spend']}"
    ))
    label_row("fcfe", "FCFE (equity)", "A$M")
    put_formula("fcfe", lambda pr, cl: (
        f"={cl}{R['ops_flag']}*({cl}{R['pat']}+{cl}{R['book_dep']}-{cl}{R['debt_repay']})-{cl}{R['equity_spend']}"
    ))
    label_row("cfads", "CFADS", "A$M")
    put_formula("cfads", lambda pr, cl: f"={cl}{R['ebitda']}")
    label_row("dscr", "DSCR", "ratio")
    put_formula("dscr", lambda pr, cl: (
        f"=IF(({cl}{R['debt_interest']}+{cl}{R['debt_repay']})>0,"
        f"{cl}{R['cfads']}/({cl}{R['debt_interest']}+{cl}{R['debt_repay']}),\"\")"
    ), "0.00")
    label_row("cum_fcfe", "Cumulative FCFE", "A$M")
    put_formula("cum_fcfe", lambda pr, cl: f"=SUM(${firstcol}{R['fcfe']}:{cl}{R['fcfe']})")

    # ── Capital & funding summary (single-value block for the Outputs sheet) ──
    cells: dict[str, str] = {}

    def summary(label: str, key: str, formula: str, fmt: str = "#,##0.0") -> None:
        nonlocal row
        r = row
        _text(ws.cell(r, 2), label).font = Font(bold=True)
        c = ws.cell(r, _pcol(0), formula)
        c.number_format = fmt
        c.fill = _PREFILL
        cells[key] = f"Model!$D${r}"
        row += 1

    label_row("sum_sec", "Capital & funding summary", section=True)
    summary("Total capital spend (nominal)", "total_capex_nominal",
            f"=SUM({firstcol}{R['capital_spend']}:{lastcol}{R['capital_spend']})")
    summary("Total interest during construction", "total_idc",
            f"=SUM({firstcol}{R['debt_idc']}:{lastcol}{R['debt_idc']})")
    summary("Total funding (incl. IDC)", "total_funding",
            f"=SUM({firstcol}{R['capital_spend']}:{lastcol}{R['capital_spend']})"
            f"+SUM({firstcol}{R['debt_idc']}:{lastcol}{R['debt_idc']})")
    summary("Debt at COD", "debt_at_cod",
            f"=SUM({firstcol}{R['debt_draw']}:{lastcol}{R['debt_draw']})"
            f"+SUM({firstcol}{R['debt_idc']}:{lastcol}{R['debt_idc']})")
    summary("Total equity", "total_equity",
            f"=SUM({firstcol}{R['equity_spend']}:{lastcol}{R['equity_spend']})")
    summary("Gearing", "gearing",
            f"={cells['debt_at_cod']}/({cells['debt_at_cod']}+{cells['total_equity']})", "0.0%")
    summary("Annual generation", "annual_gen",
            f"={E['total_solar_gwh']}+{E['total_nonsolar_gwh']}", "#,##0.0 \"GWh\"")
    summary("Fixed O&M (p.a.)", "fixed_om_total", f"={fixed_om_expr}", "#,##0.0")
    summary("Annuity factor", "annuity",
            f"=(1-(1+{I['discount_rate']})^-{I['operating_life']})/{I['discount_rate']}", "0.000")
    summary("LCOE", "lcoe",
            f"=IFERROR(({cells['total_capex_nominal']}*1000000/{cells['annuity']}"
            f"+{cells['fixed_om_total']}*1000000)/({cells['annual_gen']}*1000),\"n/a\")",
            "#,##0.0 \"A$/MWh\"")

    # remember key ranges for the Outputs sheet
    first, last = _pcol(0), _pcol(n - 1)
    fl, ll = get_column_letter(first), get_column_letter(last)
    ranges = {
        "fcff": f"Model!{fl}{R['fcff']}:{ll}{R['fcff']}",
        "fcfe": f"Model!{fl}{R['fcfe']}:{ll}{R['fcfe']}",
        "ebitda": f"Model!{fl}{R['ebitda']}:{ll}{R['ebitda']}",
        "dscr": f"Model!{fl}{R['dscr']}:{ll}{R['dscr']}",
        "cum_fcfe": f"Model!{fl}{R['cum_fcfe']}:{ll}{R['cum_fcfe']}",
        "header": f"Model!{fl}${hdr}:{ll}${hdr}",
    }
    ranges.update(cells)
    wb._fm_ranges = ranges  # type: ignore[attr-defined]
    return ranges


# ── Outputs sheet (every number a live formula) ───────────────────────────────


def _payback_formula(cum_range: str, header_range: str) -> str:
    """Fractional year of the last negative-to-positive cumulative-FCFE crossing.

    ``L`` is the year of the last negative cumulative balance; the fractional
    part interpolates to zero against the following year's balance. Mirrors
    :func:`ppa.financial_model._payback` (0-based years)."""
    last_neg = f"LOOKUP(2,1/({cum_range}<0),{header_range})"
    pos = f"MATCH({last_neg},{header_range},0)"
    frac = (f"(0-INDEX({cum_range},{pos}))"
            f"/(INDEX({cum_range},{pos}+1)-INDEX({cum_range},{pos}))")
    return f"=IFERROR({last_neg}+{frac},IFERROR({last_neg},\"n/a\"))"


def _write_outputs(
    wb: Workbook,
    result: ProjectFinanceResult,
    I: dict[str, str],
    ranges: dict[str, str],
) -> None:
    ws = wb.create_sheet("Outputs", 0)  # first sheet
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 16

    _text(ws["B1"], "Financial Model — Key Outputs")
    ws["B1"].font = _TITLE
    _text(ws["B2"], f"Scenario: {result.energy.name}")
    ws["B2"].font = Font(italic=True, color="808080")

    row = 4

    def kpi(label: str, fmt: str, formula: str) -> None:
        nonlocal row
        _text(ws.cell(row, 2), label).font = Font(bold=True)
        c = ws.cell(row, 3)
        c.value = formula
        c.number_format = fmt
        c.fill = _PREFILL
        c.border = _BORDER
        row += 1

    rate = I["discount_rate"]
    dscr = ranges["dscr"]

    kpi("Project IRR (FCFF)", "0.0%", f"=IFERROR(IRR({ranges['fcff']}),\"n/a\")")
    kpi("Equity IRR (FCFE)", "0.0%", f"=IFERROR(IRR({ranges['fcfe']}),\"n/a\")")
    kpi("NPV @ WACC (project)", "#,##0.0 \"A$M\"",
        f"=NPV({rate},{ranges['fcff']})*(1+{rate})")
    kpi("Gearing", "0.0%", f"={ranges['gearing']}")
    kpi("Total funding (incl. IDC)", "#,##0.0 \"A$M\"", f"={ranges['total_funding']}")
    kpi("Total debt", "#,##0.0 \"A$M\"", f"={ranges['debt_at_cod']}")
    kpi("Total equity", "#,##0.0 \"A$M\"", f"={ranges['total_equity']}")
    kpi("Minimum DSCR", "0.00", f"=IF(COUNT({dscr})>0,MIN({dscr}),\"n/a\")")
    kpi("Average DSCR", "0.00", f"=IF(COUNT({dscr})>0,AVERAGE({dscr}),\"n/a\")")
    kpi("Equity payback", "0.0 \"yrs\"",
        _payback_formula(ranges["cum_fcfe"], ranges["header"]))
    kpi("LCOE", "#,##0.0 \"A$/MWh\"", f"={ranges['lcoe']}")

    _text(ws.cell(row + 1, 2),
          "Every number above is a live formula referencing the Model / Energy / "
          "Hourly sheets — edit an input and the returns recompute. Debt "
          "drawdown/IDC are pre-solved (circular); re-run the toolkit to re-size "
          "debt after large cost changes.").font = (
        Font(italic=True, color="A0A0A0", size=9))


# ── Notes sheet ───────────────────────────────────────────────────────────────


def _write_notes(wb: Workbook) -> None:
    ws = wb.create_sheet("Notes")
    ws.column_dimensions["B"].width = 100
    _text(ws["B1"], "Model notes & simplifications")
    ws["B1"].font = _TITLE
    notes = [
        "This workbook is a streamlined export of the PyPSA-PPA toolkit's project-finance model.",
        "",
        "Live (formula-driven, recompute on edit):",
        "  • The single Hourly sheet holds every simulated year's hourly dispatch, stacked",
        "    vertically under a Year column; the Energy-sheet annual totals are AVERAGEs of",
        "    per-year SUMIFS roll-ups of those hours. Edit the hourly data and the totals (and",
        "    the model) follow.",
        "  • Capex & devex — per-technology build/connection/devex cost × capacity × indexation,",
        "    with per-technology subtotal rows (spend timing baked per year; edit a cost and it",
        "    flows through).",
        "  • Year 0 is the base year: indexation multipliers are 1.0 at year 0 and compound",
        "    (1 + rate)^year thereafter — no inflation is applied in year 0.",
        "  • Volumes (PPA / merchant / penalty) and average prices per MWh, then the revenue",
        "    lines computed from those rows.",
        "  • The debt corkscrew — opening balance → drawdown → IDC/interest → repayment →",
        "    closing balance — with closing feeding the next year's opening.",
        "  • Book/tax depreciation (straight-line, capped at the live asset base).",
        "  • Taxable income, loss carry-forward and income tax; EBIT / PBT / PAT / net margin.",
        "  • Simplified balance sheet (gross assets, accumulated depreciation, net debt, funds).",
        "  • Cash-flow statement (operating / investing / financing) and cumulative net cash flow.",
        "  • PBT, PAT, FCFF, FCFE, DSCR, and the Project/Equity IRR outputs.",
        "",
        "Toolkit-sized values (green) — edit to override:",
        "  • Debt drawdown, IDC and the contracted/uncontracted tranche split are circular",
        "    (debt size depends on IDC which depends on drawdown), so they are pre-solved.",
        "    Changing capex therefore updates returns but not the debt amount — re-run the",
        "    toolkit to re-size debt.",
        "",
        "Simplifications (consistent with the source model):",
        "  • No working capital, no dividends, no terminal/decommissioning value.",
        "  • Devex is a single bullet per technology, paid at FID (the `development_start`",
        "    period), immediately before construction spend begins. FID and financial close",
        "    are the same period, so devex is refinanced by debt drawn from financial close.",
        "  • One representative operating year, escalated by indexation from year 1.",
        "  • Solar hours defined as 09:00–17:00.",
    ]
    for i, line in enumerate(notes):
        _text(ws.cell(2 + i, 2), line)
